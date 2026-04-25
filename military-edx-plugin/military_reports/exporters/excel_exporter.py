"""
military_reports/exporters/excel_exporter.py

Export queryset of MilitaryUserProfile to Excel using openpyxl.
"""
import io
from datetime import date

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def export_excel(queryset) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงาน"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "ลำดับ", "ชื่อ-นามสกุล", "ชั้นยศ", "หน่วยต้นสังกัด", "หน่วยรอง",
        "อายุ (ปี)", "อายุราชการ (ปี)", "วันเริ่มรับราชการ",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, profile in enumerate(queryset, start=1):
        ws.append([
            idx,
            profile.full_name_th,
            profile.get_rank_display(),
            profile.unit,
            profile.sub_unit,
            profile.age,
            profile.service_years,
            profile.service_start_date.strftime("%d/%m/%Y"),
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"report_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
